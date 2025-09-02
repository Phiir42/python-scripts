#!/usr/bin/env python3
"""
prep_prism_cortical_layers.py
─────────────────────────────
Create a “Prism Prep” sheet from the cortical-layer workbook.
"""

# ── USER SETTINGS ──────────────────────────────────────────────────────────────
INPUT_FILE  = r"C:\Users\clchr\OneDrive - Stanford\Research Documents\AD Project\2025\AD_Lipid_Statistics_CorticalLayers.xlsx"
OUTPUT_FILE = None          # None ➜ “…_prism.xlsx” alongside the input file
# ───────────────────────────────────────────────────────────────────────────────

import pathlib
from collections import defaultdict
from statistics import mean

import openpyxl as oxl
from openpyxl.styles import Alignment, Font

# ── CONSTANTS ─────────────────────────────────────────────────────────────────
CELL_TYPES  = ["Microglia", "Astrocytes", "Neurons"]
CONDITIONS  = ["Control", "AD33", "AD44"]

LAYERS   = ["Layer I", "Layer II", "Layer III",
            "Layer IV", "Layer V", "Layer VI", "White Matter"]

METRICS  = ["Lipids", "Lipofuscin", "Lipidated Lipofuscin", "Myelination", "Amyloid"]

HEADER_ROWS = 2                     # rows 1–2 = column headers
START_ROW   = HEADER_ROWS + 1       # first numeric row (1-based index)

# ──────────────────────────────────────────────────────────────────────────────
#  STEP 1  ─ Collect per-replicate means from every worksheet
# ──────────────────────────────────────────────────────────────────────────────
data = {m: {l: defaultdict(lambda: defaultdict(list)) for l in LAYERS}
        for m in METRICS}
rep_counts = {ct: {cond: 0 for cond in CONDITIONS} for ct in CELL_TYPES}
donor_blocks = {ct: {cond: [] for cond in CONDITIONS} for ct in CELL_TYPES}

wb_in = oxl.load_workbook(INPUT_FILE, data_only=True)

for cell_type in CELL_TYPES:
    for cond in CONDITIONS:
        ws      = wb_in[f"{cond} {cell_type}"]

        # locate the 3×7 data columns
        col_map, current_layer = {}, None
        for col in range(2, ws.max_column + 1):
            hdr_layer  = ws.cell(row=1, column=col).value
            hdr_metric = ws.cell(row=2, column=col).value
            if hdr_layer:
                current_layer = hdr_layer.strip()
            if current_layer in LAYERS and hdr_metric in METRICS:
                col_map[(current_layer, hdr_metric.strip())] = col

        # ── replicate blocks in column A  ( **filter out header rows** ) ─────
        replicate_ranges = sorted(
            [
                rng for rng in ws.merged_cells.ranges
                if (
                    rng.min_col == rng.max_col == 1         # only column A
                    and rng.min_row >= START_ROW            # skip header merge
                )
            ],
            key=lambda r: r.min_row
        )
        rep_counts[cell_type][cond] = len(replicate_ranges)
        
        for rng in replicate_ranges:
            file_name_in_A = ws.cell(rng.min_row, 1).value
            donor_id = file_name_in_A.split("-")[1]
            donor_blocks[cell_type][cond].append(donor_id)

        # average each replicate block
        for rng in replicate_ranges:                       # keep top→bottom
            rows = range(rng.min_row, rng.max_row + 1)
            for layer in LAYERS:
                for metric in METRICS:
                    col  = col_map[(layer, metric)]
                    vals = [
                        ws.cell(r, col).value
                        for r in rows
                        if isinstance(ws.cell(r, col).value, (int, float))
                    ]
                    avg = mean(vals) if vals else None
                    data[metric][layer][cell_type][cond].append(avg)

wb_in.close()

# ──────────────────────────────────────────────────────────────────────────────
#  STEP 2  ─ Build the “Prism Prep” sheet
# ──────────────────────────────────────────────────────────────────────────────
wb_out = oxl.load_workbook(INPUT_FILE)
if "Prism Prep" in wb_out.sheetnames:
    del wb_out["Prism Prep"]
ws_out = wb_out.create_sheet("Prism Prep")

# header rows (1-3)
cur_col = 3                           # col A = Metric, col B = Layer
cell_merges, cond_merges = [], []

for cell_type in CELL_TYPES:
    ct_start = cur_col
    for cond in CONDITIONS:
        reps      = rep_counts[cell_type][cond]
        cond_start = cur_col
        for r in range(1, reps + 1):
            ws_out.cell(row=2, column=cur_col, value=cond)
            ws_out.cell(row=3, column=cur_col, value=f"R{r}")
            cur_col += 1
        cond_merges.append((2, cond_start, cur_col - 1))
    cell_merges.append((1, ct_start, cur_col - 1))
    ws_out.cell(row=1, column=ct_start, value=cell_type)

for r, c1, c2 in cell_merges:
    if c2 > c1:
        ws_out.merge_cells(start_row=r, end_row=r, start_column=c1, end_column=c2)
for r, c1, c2 in cond_merges:
    if c2 > c1:
        ws_out.merge_cells(start_row=r, end_row=r, start_column=c1, end_column=c2)

# data rows
row_ptr = 4
for metric in METRICS:
    metric_start = row_ptr
    for layer in LAYERS:
        ws_out.cell(row=row_ptr, column=2, value=layer)
        col_ptr = 3
        for cell_type in CELL_TYPES:
            for cond in CONDITIONS:
                for val in data[metric][layer][cell_type][cond]:
                    ws_out.cell(row=row_ptr, column=col_ptr, value=val)
                    col_ptr += 1
        row_ptr += 1
    ws_out.cell(row=metric_start, column=1, value=metric)
    ws_out.merge_cells(start_row=metric_start,
                       end_row=row_ptr - 1,
                       start_column=1, end_column=1)

ws_out.freeze_panes = "C4"

# save
in_path  = pathlib.Path(INPUT_FILE).expanduser().resolve()
out_path = (pathlib.Path(OUTPUT_FILE).expanduser().resolve()
            if OUTPUT_FILE
            else in_path.with_name(in_path.stem + "_prism.xlsx"))

wb_out.save(out_path)
print(f"✅  Prism Prep sheet added.  Saved to: {out_path}")

# ──────────────────────────────────────────────────────────────────────────────
# STEP 3 ─ Build “Prism Prep – Donor” for Myelination & Amyloid combined
# ──────────────────────────────────────────────────────────────────────────────

# 3a) Build donors_per_condition exactly as above
donors_per_condition = {}
for cond in CONDITIONS:
    combined = []
    for ct in CELL_TYPES:
        combined += donor_blocks[ct][cond]
    seen = set()
    unique_list = []
    for d in combined:
        if d not in seen:
            seen.add(d)
            unique_list.append(d)
    donors_per_condition[cond] = unique_list

# 3b) Compute one combined value per donor ↔ layer ↔ metric
combined_data = {
    m: {layer: {cond: {} for cond in CONDITIONS} for layer in LAYERS}
    for m in ["Myelination", "Amyloid"]
}

for metric in ["Myelination", "Amyloid"]:
    for layer in LAYERS:
        for cond in CONDITIONS:
            for donor_id in donors_per_condition[cond]:
                vals_for_donor = []
                for ct in CELL_TYPES:
                    block_list = donor_blocks[ct][cond]
                    data_list  = data[metric][layer][ct][cond]
                    for idx, block_donor in enumerate(block_list):
                        if block_donor == donor_id:
                            v = data_list[idx]
                            # Only append if not None:
                            if v is not None:
                                vals_for_donor.append(v)

                # Now compute mean only if there’s at least one numeric value
                if vals_for_donor:
                    combined_data[metric][layer][cond][donor_id] = mean(vals_for_donor)
                else:
                    combined_data[metric][layer][cond][donor_id] = None

# 3c) Create the new "Prism Prep - Donor" sheet
sheet_name = "Prism Prep - Donor"
if sheet_name in wb_out.sheetnames:
    del wb_out[sheet_name]
ws_donor = wb_out.create_sheet(sheet_name)

# Header rows (1–2): Metric | Layer | [Condition1 merged] [Condition2 merged] ...
ws_donor.cell(row=1, column=1, value="Metric")
ws_donor.cell(row=1, column=2, value="Layer")
ws_donor.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
ws_donor.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
ws_donor.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws_donor.cell(row=1, column=2).alignment = Alignment(horizontal='center', vertical='center')
ws_donor.cell(row=1, column=1).font = Font(bold=True)
ws_donor.cell(row=1, column=2).font = Font(bold=True)

cur_col = 3
for cond in CONDITIONS:
    n_donors = len(donors_per_condition[cond])
    if n_donors == 0:
        continue
    start = cur_col
    ws_donor.merge_cells(
        start_row=1,
        start_column=start,
        end_row=1,
        end_column=start + n_donors - 1
    )
    ws_donor.cell(row=1, column=start, value=cond).alignment = Alignment(horizontal='center')
    ws_donor.cell(row=1, column=start).font = Font(bold=True)
    for r in range(1, n_donors + 1):
        ws_donor.cell(row=2, column=cur_col, value=f"R{r}").alignment = Alignment(horizontal='center')
        ws_donor.cell(row=2, column=cur_col).font = Font(bold=True)
        cur_col += 1

# 3d) Fill rows 3→, one block for Myelination then one for Amyloid
row_ptr = 3
for metric in ["Myelination", "Amyloid"]:
    block_start = row_ptr
    for layer in LAYERS:
        ws_donor.cell(row=row_ptr, column=2, value=layer)
        col_ptr = 3
        for cond in CONDITIONS:
            for donor_id in donors_per_condition[cond]:
                ws_donor.cell(
                    row=row_ptr,
                    column=col_ptr,
                    value=combined_data[metric][layer][cond][donor_id]
                )
                col_ptr += 1
        row_ptr += 1
    # Merge the metric label down its 7‐row block
    ws_donor.cell(row=block_start, column=1, value=metric)
    ws_donor.merge_cells(
        start_row=block_start,
        end_row=row_ptr - 1,
        start_column=1,
        end_column=1
    )
    ws_donor.cell(row=block_start, column=1).alignment = Alignment(horizontal='center', vertical='center')

ws_donor.freeze_panes = "C4"

# 3e) Save the workbook one more time
wb_out.save(out_path)
print(f"✅  Prism Prep - Donor sheet added.  Saved to: {out_path}")
