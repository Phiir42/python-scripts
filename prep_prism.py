#!/usr/bin/env python3
"""
prep_prism.py — run inside Spyder (or any IDE)

Creates / replaces the following worksheets:

    1. Prism Prep                     (raw, zeros kept)
    2. Prism Averages                 (donor-averaged, zeros kept)
    3. Prism Prep (no zeros)          (raw, zeros removed)
    4. Prism Averages (no zeros)      (donor-averaged, zeros removed)
    5. XY {Condition} {CellType}      (myelination vs metrics)
"""

# ── USER SETTINGS ─────────────────────────────────────────────────────────────
INPUT_FILE  = r"C:\Users\clchr\OneDrive - Stanford\Research Documents\AD Project\2025\AD Lipid Statistics.xlsx"   # ← edit to your aggregated file
OUTPUT_FILE = None        # leave None to get “…_prism.xlsx” in same folder
# ──────────────────────────────────────────────────────────────────────────────

import pathlib
import pandas as pd
from openpyxl import load_workbook

# ── CONSTANTS ────────────────────────────────────────────────────────────────
CELL_TYPES = ["Microglia", "Astrocytes", "Neurons"]
METRICS = [
    ("pure_lipid_percentage",       "Lipid Areas"),
    ("lipofuscin_percentage",       "Lipofuscin Areas"),
    ("lipid_lipofuscin_percentage", "Lipidated Lipofuscin Areas"),
    ("myelination_percentage",      "Myelination Percentage"),
]
CONDITIONS = ["Control", "AD33", "AD44"]
# ──────────────────────────────────────────────────────────────────────────────


def _cond_donor_stub(file_name: str) -> str:
    parts = file_name.split("-", 2)
    return "-".join(parts[:2]) if len(parts) >= 2 else file_name


def build_prism_dataframe(xlsx: pathlib.Path, *, remove_zeros: bool = False) -> pd.DataFrame:
    cols, max_len = [], 0
    for metric_key, _ in METRICS:
        for cell in CELL_TYPES:
            for cond in CONDITIONS:
                sheet = f"{cond} {cell}"
                df_sheet = pd.read_excel(xlsx, sheet_name=sheet)
                if metric_key in df_sheet.columns:
                    s = df_sheet[metric_key]
                else:
                    s = pd.Series([pd.NA] * len(df_sheet), name=metric_key)
                if remove_zeros:
                    s = s[s != 0].reset_index(drop=True)
                cols.append(s)
                max_len = max(max_len, len(s))

    padded = [s.reindex(range(max_len)).reset_index(drop=True) for s in cols]
    return pd.concat(padded, axis=1)


def build_prism_average_dataframe(xlsx: pathlib.Path, *, remove_zeros: bool = False) -> pd.DataFrame:
    series_list = []
    donors = set()
    for metric_key, _ in METRICS:
        for cell in CELL_TYPES:
            for cond in CONDITIONS:
                sheet = f"{cond} {cell}"
                df_full = pd.read_excel(xlsx, sheet_name=sheet)
                if metric_key in df_full.columns:
                    df = df_full[["file_name", metric_key]].copy()
                else:
                    df = df_full[["file_name"]].copy()
                    df[metric_key] = pd.NA
                if remove_zeros:
                    df.loc[df[metric_key] == 0, metric_key] = pd.NA
                df["cond_donor"] = df["file_name"].map(_cond_donor_stub)
                s = df.groupby("cond_donor", dropna=False)[metric_key].mean()
                series_list.append(s)
                donors.update(s.index)

    donor_list = sorted(donors)
    padded = [s.reindex(donor_list) for s in series_list]
    wide = pd.concat(padded, axis=1)
    wide.insert(0, "cond_donor", donor_list)
    return wide


def _write_headers(ws, start_col: int):
    col = start_col
    for _metric_key, metric_title in METRICS:
        for cell in CELL_TYPES:
            start, end = col, col + len(CONDITIONS) - 1
            ws.merge_cells(start_row=1, start_column=start,
                           end_row=1,   end_column=end)
            ws.cell(row=1, column=start).value = f"{cell} {metric_title}"
            col = end + 1
    col = start_col
    for _ in range(len(METRICS) * len(CELL_TYPES)):
        for cond in CONDITIONS:
            ws.cell(row=2, column=col, value=cond)
            col += 1


def _write_df(ws, df: pd.DataFrame, *, id_col: bool = False):
    offset = 1 if id_col else 0
    if id_col:
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.cell(row=1, column=1, value="Donor (cond-stub)")
    for r, row in enumerate(df.itertuples(index=False), start=3):
        for c, val in enumerate(row, start=1+offset):
            ws.cell(row=r, column=c, value=val)


def add_prism_sheets(infile: pathlib.Path, outfile: pathlib.Path) -> None:
    wb = load_workbook(infile)
    df_raw    = build_prism_dataframe(infile, remove_zeros=False)
    df_raw_nz = build_prism_dataframe(infile, remove_zeros=True)
    df_avg    = build_prism_average_dataframe(infile, remove_zeros=False)
    df_avg_nz = build_prism_average_dataframe(infile, remove_zeros=True)

    sheet_specs = [
        ("Prism Prep",                    df_raw,     False),
        ("Prism Averages",                df_avg,     True ),
        ("Prism Prep (no zeros)",         df_raw_nz,  False),
        ("Prism Averages (no zeros)",     df_avg_nz,  True ),
    ]
    for name, df, has_id in sheet_specs:
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        _write_headers(ws, start_col=2 if has_id else 1)
        _write_df(ws, df, id_col=has_id)

    # ── NEW: add XY scatter data sheets ─────────────────────────────────────
    metrics_to_plot = [
        "pure_lipid_percentage",
        "lipofuscin_percentage",
        "lipid_lipofuscin_percentage",
    ]
    for cond in CONDITIONS:
        for cell in CELL_TYPES:
            sheet_in = f"{cond} {cell}"
            if sheet_in not in wb.sheetnames:
                continue
            # load the detailed sheet to get raw columns
            df_sheet = pd.read_excel(infile, sheet_name=sheet_in)
            cols = ["myelination_percentage"] + metrics_to_plot
            existing = [c for c in cols if c in df_sheet.columns]
            if "myelination_percentage" not in existing:
                continue
            xy_df = df_sheet[existing].dropna(subset=["myelination_percentage"])
            xy_df = xy_df.sort_values(by="myelination_percentage")
            xy_name = f"XY {cond} {cell}"
            if xy_name in wb.sheetnames:
                del wb[xy_name]
            ws_xy = wb.create_sheet(xy_name)
            # write header row
            for idx, col_name in enumerate(existing, start=1):
                ws_xy.cell(row=1, column=idx, value=col_name)
            # write data rows
            for r_idx, row in enumerate(xy_df.itertuples(index=False), start=2):
                for c_idx, val in enumerate(row, start=1):
                    ws_xy.cell(row=r_idx, column=c_idx, value=val)

    wb.save(outfile)


def main() -> None:
    in_path = pathlib.Path(INPUT_FILE).expanduser().resolve()
    if not in_path.is_file():
        raise FileNotFoundError(f"Input file not found: {in_path}")
    out_path = (
        pathlib.Path(OUTPUT_FILE).expanduser().resolve()
        if OUTPUT_FILE
        else in_path.with_name(in_path.stem + "_prism.xlsx")
    )
    add_prism_sheets(in_path, out_path)
    print(f"✅  All Prism sheets (including XY scatter) added.  Saved to: {out_path}")


if __name__ == "__main__":
    main()
