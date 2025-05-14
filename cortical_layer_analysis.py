import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

def determine_layer(file_name, z_stack):
    """
    Determines which layer (1..7) a given z_stack belongs to,
    based on the skip logic you provided.
    """
    normal_files = {
        "Control-S0536-DAPI-IBA1-GFAP-100X-StacksMicroglia.nd2",
        "Control-S0536-DAPI-IBA1-GFAP-100X-StacksAstrocytes.nd2",
        "AD33-S2143-DAPI-IBA1-GFAP-100X-StacksMicroglia.nd2",
        "AD33-S2143-DAPI-IBA1-GFAP-100X-StacksAstrocytes.nd2",
        "AD44-S1342-DAPI-IBA1-GFAP-100X-StacksMicroglia.nd2",
        "AD44-S1342-DAPI-IBA1-GFAP-100X-StacksAstrocytes.nd2",
        "Control-S0536-DAPI-TUJ-LAMP2-100X-StacksNeurons.nd2",
        "AD33-S2146-DAPI-IBA1-GFAP-100X-StacksMicroglia.nd2",
        "AD33-S2146-DAPI-IBA1-GFAP-100X-StacksAstrocytes.nd2",
        "AD44-S1563-DAPI-IBA1-GFAP-100X-StacksMicroglia.nd2",
        "AD44-S1563-DAPI-IBA1-GFAP-100X-StacksAstrocytes.nd2",
        "Control-S2302-DAPI-IBA1-GFAP-100X-StacksMicroglia.nd2",
        "Control-S2302-DAPI-IBA1-GFAP-100X-StacksAstrocytes.nd2",
        "AD33-S2146-DAPI-TUJ-LAMP2-100X-StacksNeurons.nd2",
        "AD44-S1563-DAPI-TUJ-LAMP2-100X-StacksNeurons.nd2",
        "Control-S2302-DAPI-TUJ-LAMP2-100X-StacksNeurons.nd2"
    }

    ad44_tuj_lamp2 = "AD44-S1342-DAPI-TUJ-LAMP2-100X-StacksNeurons.nd2"
    ad33_tuj_lamp2 = "AD33-S2143-DAPI-TUJ-LAMP2-100X-StacksNeurons.nd2"
    control_s2218  = "Control-S2218-DAPI-TUJ-LAMP2-100X-StacksNeurons.nd2"

    if file_name in normal_files:
        # 1..6 => Layer I, 7..12 => Layer II, etc.
        if 1 <= z_stack <= 6:
            return 1
        elif 7 <= z_stack <= 12:
            return 2
        elif 13 <= z_stack <= 18:
            return 3
        elif 19 <= z_stack <= 24:
            return 4
        elif 25 <= z_stack <= 30:
            return 5
        elif 31 <= z_stack <= 36:
            return 6
        elif 37 <= z_stack <= 42:
            return 7
        else:
            return None

    elif file_name == ad44_tuj_lamp2:
        # skip Layer I
        if 1 <= z_stack <= 6:
            return 2
        elif 7 <= z_stack <= 12:
            return 3
        elif 13 <= z_stack <= 18:
            return 4
        elif 19 <= z_stack <= 24:
            return 5
        elif 25 <= z_stack <= 30:
            return 6
        elif 31 <= z_stack <= 36:
            return 7
        else:
            return None

    elif file_name == ad33_tuj_lamp2:
        # skip I, II, White Matter
        if 1 <= z_stack <= 6:
            return 3
        elif 7 <= z_stack <= 12:
            return 4
        elif 13 <= z_stack <= 18:
            return 5
        elif 19 <= z_stack <= 24:
            return 6
        else:
            return None

    elif file_name == control_s2218:
        # skip I, II, III
        if 1 <= z_stack <= 6:
            return 4
        elif 7 <= z_stack <= 12:
            return 5
        elif 13 <= z_stack <= 18:
            return 6
        elif 19 <= z_stack <= 24:
            return 7
        else:
            return None

    else:
        return None

def create_formatted_worksheet(wb, sheet_name):
    """
    Create a worksheet with merged headers:
      Row 1 => "Layer I" ... "White Matter" (7 merges)
      Row 2 => "Lipids", "Lipidated Lipofuscin", "Lipofuscin" subcolumns
    Column A => 'file_name' (merged vertically in each file_name block).
    """
    ws = wb.create_sheet(title=sheet_name)

    # Create top-left cell for "file_name", spanning row 1..2 horizontally
    # but we'll only truly finalize merging that column for data blocks later.
    ws.cell(row=1, column=1, value="file_name")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=1, column=1).font = Font(bold=True)

    # 7 layers, each has 3 subcolumns:
    layer_names = ["Layer I", "Layer II", "Layer III", "Layer IV", "Layer V", "Layer VI", "White Matter"]
    subcolumns = ["Lipids", "Lipidated Lipofuscin", "Lipofuscin"]

    # For each layer, we merge 3 columns in row 1
    # Then row 2 has the subcolumn names.
    start_col = 2  # columns start from B for the first layer
    for layer_name in layer_names:
        # Merge row 1 over these 3 columns
        ws.merge_cells(
            start_row=1,
            start_column=start_col,
            end_row=1,
            end_column=start_col + 2
        )
        ws.cell(row=1, column=start_col, value=layer_name)
        ws.cell(row=1, column=start_col).font = Font(bold=True)
        ws.cell(row=1, column=start_col).alignment = Alignment(horizontal='center', vertical='center')

        # Row 2 => subcol names
        for j, subcol_name in enumerate(subcolumns):
            c = start_col + j
            ws.cell(row=2, column=c, value=subcol_name)
            ws.cell(row=2, column=c).font = Font(bold=True)
            ws.cell(row=2, column=c).alignment = Alignment(horizontal='center', vertical='center')

        start_col += 3

    return ws

def main():
    input_filename = "AD Lipid Statistics.xlsx"
    output_filename = "AD_Lipid_Statistics_Reformatted.xlsx"

    xl = pd.ExcelFile(input_filename)
    sheet_names = xl.sheet_names

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name in sheet_names:
        df = pd.read_excel(xl, sheet_name=sheet_name)

        # Check columns
        required_cols = [
            'file_name',
            'z_stack',
            'pure_lipid_percentage',
            'lipid_lipofuscin_percentage',
            'lipofuscin_percentage'
        ]
        if not all(col in df.columns for col in required_cols):
            print(f"Skipping '{sheet_name}' - missing required columns.")
            continue

        # Create the new sheet with the 2-row headers
        ws = create_formatted_worksheet(wb, sheet_name)

        # Group the data by file_name -> layer -> list of (lipids, lipid_lipo, lipo)
        grouped_data = {}
        for i, row_data in df.iterrows():
            fn = row_data['file_name']
            zs = row_data['z_stack']
            pl = row_data['pure_lipid_percentage']
            llf = row_data['lipid_lipofuscin_percentage']
            lpf = row_data['lipofuscin_percentage']

            layer = determine_layer(fn, zs)
            if layer is None:
                continue

            # Store in grouped_data
            if fn not in grouped_data:
                grouped_data[fn] = {i: [] for i in range(1,8)}
            grouped_data[fn][layer].append((pl, llf, lpf))

        # Now we’ll write the data. Start from row=3 (below the headers).
        current_row = 3

        # We have 7 columns of interest (Layer 1..7),
        # each taking up 3 subcolumns => total columns from 2..22
        def layer_to_cols(layer):
            # layer 1 => columns 2,3,4
            # layer 2 => columns 5,6,7
            # ...
            start_col = 2 + (layer - 1)*3
            return start_col, start_col+1, start_col+2

        # Write each file_name in a block
        for file_name, layer_dict in grouped_data.items():

            # For each of the 7 layers, we have some number of data points
            # The block height = the maximum number of data points among all layers
            max_points = max(len(layer_dict[l]) for l in range(1,8))

            # If a file_name has 0 data points altogether, skip
            if max_points == 0:
                continue

            # We will fill rows current_row..(current_row+max_points-1)
            # and place data in the correct columns for each layer
            block_start = current_row
            block_end   = current_row + max_points - 1

            # (Optional) Put the file_name in col=1, merged from block_start..block_end
            ws.cell(row=block_start, column=1, value=file_name)
            if max_points > 1:
                # vertically merge that cell down
                ws.merge_cells(
                    start_row=block_start,
                    start_column=1,
                    end_row=block_end,
                    end_column=1
                )
            # style it
            ws.cell(row=block_start, column=1).alignment = Alignment(
                horizontal='center',
                vertical='center'
            )
            ws.cell(row=block_start, column=1).font = Font(bold=False)

            # Now fill row by row
            # For each i in 0..(max_points-1), place data for each layer in that row
            for i in range(max_points):
                # The row we’re writing to:
                row_i = block_start + i

                # For each layer 1..7, check if we have a data point
                for lyr in range(1, 8):
                    data_points = layer_dict[lyr]  # list of (pl, llf, lpf)
                    # If i < len(data_points), we place it, otherwise leave blank
                    if i < len(data_points):
                        pl, llf, lpf = data_points[i]
                        c1, c2, c3 = layer_to_cols(lyr)
                        ws.cell(row=row_i, column=c1, value=pl)
                        ws.cell(row=row_i, column=c2, value=llf)
                        ws.cell(row=row_i, column=c3, value=lpf)
                    else:
                        # This layer doesn't have an (i+1)-th data point => blank
                        pass

            # After filling these max_points rows for this file_name,
            # move current_row down for the next file_name block
            current_row = block_end + 1

    # Finally, save the new workbook
    wb.save(output_filename)
    print(f"Saved '{output_filename}' with your desired layout.")

if __name__ == "__main__":
    main()
