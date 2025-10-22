#!/usr/bin/env python3
"""
prep_prism_cortical_layers.py
─────────────────────────────
Create a “Prism Prep” sheet from the cortical-layer workbook.
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from statistics import mean
from typing import DefaultDict

import openpyxl as oxl
from openpyxl.styles import Alignment, Font

# ── USER SETTINGS ──────────────────────────────────────────────────────────────
INPUT_FILE: str = (
    r"D:\OneDrive - Stanford\Research Documents\AD Project\2025\AD_Lipid_Statistics_CorticalLayers.xlsx"
)
OUTPUT_FILE: str | None = None  # None ➜ “…_prism.xlsx” alongside the input file
# ───────────────────────────────────────────────────────────────────────────────

# ── CONSTANTS ─────────────────────────────────────────────────────────────────
CELL_TYPES: list[str] = ["Microglia", "Astrocytes", "Neurons"]
CONDITIONS: list[str] = ["Control", "AD33", "AD44"]

LAYERS: list[str] = [
    "Layer I",
    "Layer II",
    "Layer III",
    "Layer IV",
    "Layer V",
    "Layer VI",
    "White Matter",
]

METRICS: list[str] = [
    "Lipids",
    "Lipofuscin",
    "Lipidated Lipofuscin",
    "Myelination"
]

HEADER_ROWS = 2  # rows 1–2 = column headers
START_ROW = HEADER_ROWS + 1  # first numeric row (1-based index)

# ── TYPES ─────────────────────────────────────────────────────────────────────
# data[metric][layer][cell_type][condition] -> list[float|None]
DataDict = dict[
    str,
    dict[str, DefaultDict[str, DefaultDict[str, list[float | None]]]],
]

# donor_blocks[cell_type][condition] -> list[str]
DonorBlocks = dict[str, dict[str, list[str]]]

# combined_data[metric][layer][condition][donor_id] -> float|None
CombinedData = dict[str, dict[str, dict[str, dict[str, float | None]]]]

# ──────────────────────────────────────────────────────────────────────────────
#  STEP 1  ─ Collect per-replicate means from every worksheet
# ──────────────────────────────────────────────────────────────────────────────
data: DataDict = {
    m: {layer: defaultdict(lambda: defaultdict(list)) for layer in LAYERS}
    for m in METRICS
}

rep_counts: dict[str, dict[str, int]] = {
    ct: {cond: 0 for cond in CONDITIONS} for ct in CELL_TYPES
}

donor_blocks: DonorBlocks = {
    ct: {cond: [] for cond in CONDITIONS} for ct in CELL_TYPES
}

wb_in = oxl.load_workbook(INPUT_FILE, data_only=True)

for cell_type in CELL_TYPES:
    for cond in CONDITIONS:
        sheet_name = f"{cond} {cell_type}"
        try:
            ws = wb_in[sheet_name]
        except KeyError:
            print(f"[WARN] Missing sheet: {sheet_name}")
            continue

        # locate the 3×7 data columns
        col_map: dict[tuple[str, str], int] = {}
        current_layer: str | None = None
        for col in range(2, ws.max_column + 1):
            hdr_layer = ws.cell(row=1, column=col).value
            hdr_metric = ws.cell(row=2, column=col).value
            if hdr_layer:
                current_layer = str(hdr_layer).strip()
            if current_layer in LAYERS and hdr_metric in METRICS:
                col_map[(current_layer, str(hdr_metric).strip())] = col

        # ── replicate blocks in column A  (filter out header rows) ─────
        replicate_ranges = sorted(
            [
                rng
                for rng in ws.merged_cells.ranges
                if (
                    rng.min_col == rng.max_col == 1  # only column A
                    and rng.min_row >= START_ROW     # skip header merge
                )
            ],
            key=lambda r: r.min_row,
        )
        if not replicate_ranges:
            print(f"[INFO] No merged replicate blocks found in column A on sheet {sheet_name}")
        rep_counts[cell_type][cond] = len(replicate_ranges)

        for rng in replicate_ranges:
            file_name_in_A = ws.cell(rng.min_row, 1).value
            parts = str(file_name_in_A).split("-")
            donor_id = parts[1] if len(parts) >= 2 else str(file_name_in_A)
            donor_blocks[cell_type][cond].append(donor_id)

        # average each replicate block
        for rng in replicate_ranges:  # keep top→bottom
            rows = range(rng.min_row, rng.max_row + 1)
            for layer in LAYERS:
                for metric in METRICS:
                    col = col_map.get((layer, metric))
                    if col is None:
                        print(f"[WARN] Missing header pair ({layer}, {metric}) in sheet {sheet_name}")
                        continue
                    
                    vals = [
                        ws.cell(r, col).value
                        for r in rows
                        if isinstance(ws.cell(r, col).value, (int, float))
                    ]
                    avg: float | None = mean(vals) if vals else None
                    data[metric][layer][cell_type][cond].append(avg)

wb_in.close()

# ──────────────────────────────────────────────────────────────────────────────
#  STEP 2  ─ Build the “Prism Prep” sheet
# ──────────────────────────────────────────────────────────────────────────────
wb_out = oxl.load_workbook(INPUT_FILE)
if "Prism Prep" in wb_out.sheetnames:
    del wb_out["Prism Prep"]
ws_out = wb_out.create_sheet("Prism Prep")

# header rows (1-3)
cur_col = 3  # col A = Metric, col B = Layer
cell_merges: list[tuple[int, int, int]] = []
cond_merges: list[tuple[int, int, int]] = []

for cell_type in CELL_TYPES:
    ct_start = cur_col
    for cond in CONDITIONS:
        reps = rep_counts[cell_type][cond]
        cond_start = cur_col
        for r in range(1, reps + 1):
            ws_out.cell(row=2, column=cur_col, value=cond)
            ws_out.cell(row=3, column=cur_col, value=f"R{r}")
            cur_col += 1
        cond_merges.append((2, cond_start, cur_col - 1))
    cell_merges.append((1, ct_start, cur_col - 1))
    ws_out.cell(row=1, column=ct_start, value=cell_type)

for r, c1, c2 in cell_merges:
    if c2 > c1:
        ws_out.merge_cells(
            start_row=r, end_row=r, start_column=c1, end_column=c2
        )
for r, c1, c2 in cond_merges:
    if c2 > c1:
        ws_out.merge_cells(
            start_row=r, end_row=r, start_column=c1, end_column=c2
        )

# data rows
row_ptr = 4
for metric in METRICS:
    metric_start = row_ptr
    for layer in LAYERS:
        ws_out.cell(row=row_ptr, column=2, value=layer)
        col_ptr = 3
        for cell_type in CELL_TYPES:
            for cond in CONDITIONS:
                for val in data[metric][layer][cell_type][cond]:
                    ws_out.cell(row=row_ptr, column=col_ptr, value=val)
                    col_ptr += 1
        row_ptr += 1
    ws_out.cell(row=metric_start, column=1, value=metric)
    ws_out.merge_cells(
        start_row=metric_start,
        end_row=row_ptr - 1,
        start_column=1,
        end_column=1,
    )

ws_out.freeze_panes = "C4"

# save
in_path = Path(INPUT_FILE).expanduser().resolve()
out_path = (
    Path(OUTPUT_FILE).expanduser().resolve()
    if OUTPUT_FILE
    else in_path.with_name(in_path.stem + "_prism.xlsx")
)

wb_out.save(out_path)
print(f"✅  Prism Prep sheet added.  Saved to: {out_path}")

# ──────────────────────────────────────────────────────────────────────────────
# STEP 3 ─ Build “Prism Prep – Donor” for Myelination
# ──────────────────────────────────────────────────────────────────────────────

# 3a) Build donors_per_condition exactly as above
donors_per_condition: dict[str, list[str]] = {}
for cond in CONDITIONS:
    combined: list[str] = []
    for ct in CELL_TYPES:
        combined += donor_blocks[ct][cond]
    seen: set[str] = set()
    unique_list: list[str] = []
    for d in combined:
        if d not in seen:
            seen.add(d)
            unique_list.append(d)
    donors_per_condition[cond] = unique_list

# 3b) Compute one combined value per donor ↔ layer ↔ metric
combined_data: CombinedData = {
    m: {layer: {cond: {} for cond in CONDITIONS} for layer in LAYERS}
    for m in ["Myelination"]
}

for metric in ["Myelination"]:
    for layer in LAYERS:
        for cond in CONDITIONS:
            for donor_id in donors_per_condition[cond]:
                vals_for_donor: list[float] = []
                for ct in CELL_TYPES:
                    block_list = donor_blocks[ct][cond]
                    data_list = data[metric][layer][ct][cond]
                    for idx, block_donor in enumerate(block_list):
                        if block_donor == donor_id:
                            v = data_list[idx]
                            if v is not None:
                                vals_for_donor.append(v)

                combined_data[metric][layer][cond][donor_id] = (
                    mean(vals_for_donor) if vals_for_donor else None
                )

# 3c) Create the new "Prism Prep - Donor" sheet
sheet_name = "Prism Prep - Donor"
if sheet_name in wb_out.sheetnames:
    del wb_out[sheet_name]
ws_donor = wb_out.create_sheet(sheet_name)

# Header rows (1–2): Metric | Layer | [Condition1 merged] [Condition2 merged] ...
ws_donor.cell(row=1, column=1, value="Metric")
ws_donor.cell(row=1, column=2, value="Layer")
ws_donor.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
ws_donor.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
ws_donor.cell(row=1, column=1).alignment = Alignment(
    horizontal="center", vertical="center"
)
ws_donor.cell(row=1, column=2).alignment = Alignment(
    horizontal="center", vertical="center"
)
ws_donor.cell(row=1, column=1).font = Font(bold=True)
ws_donor.cell(row=1, column=2).font = Font(bold=True)

cur_col = 3
for cond in CONDITIONS:
    n_donors = len(donors_per_condition[cond])
    if n_donors == 0:
        continue
    start = cur_col
    ws_donor.merge_cells(
        start_row=1,
        start_column=start,
        end_row=1,
        end_column=start + n_donors - 1,
    )
    ws_donor.cell(row=1, column=start, value=cond).alignment = Alignment(
        horizontal="center"
    )
    ws_donor.cell(row=1, column=start).font = Font(bold=True)
    for r in range(1, n_donors + 1):
        ws_donor.cell(row=2, column=cur_col, value=f"R{r}").alignment = Alignment(
            horizontal="center"
        )
        ws_donor.cell(row=2, column=cur_col).font = Font(bold=True)
        cur_col += 1

# 3d) Fill rows 3→, one block for Myelination then one for Amyloid
row_ptr = 3
for metric in ["Myelination"]:
    block_start = row_ptr
    for layer in LAYERS:
        ws_donor.cell(row=row_ptr, column=2, value=layer)
        col_ptr = 3
        for cond in CONDITIONS:
            for donor_id in donors_per_condition[cond]:
                ws_donor.cell(
                    row=row_ptr,
                    column=col_ptr,
                    value=combined_data[metric][layer][cond][donor_id],
                )
                col_ptr += 1
        row_ptr += 1
    # Merge the metric label down its 7‐row block
    ws_donor.cell(row=block_start, column=1, value=metric)
    ws_donor.merge_cells(
        start_row=block_start,
        end_row=row_ptr - 1,
        start_column=1,
        end_column=1,
    )
    ws_donor.cell(row=block_start, column=1).alignment = Alignment(
        horizontal="center", vertical="center"
    )

ws_donor.freeze_panes = "C4"

# 4) Build “Prism Prep - All” (Myelination only; RAW per-row cells; pooled across cell types; NO averaging)
SHEET_ALL = "Prism Prep - All"

# 4a) Drop if exists and create fresh
if SHEET_ALL in wb_out.sheetnames:
    del wb_out[SHEET_ALL]
ws_all = wb_out.create_sheet(SHEET_ALL)

# 4b) Re-open the INPUT workbook to read RAW cells from the 9 input sheets
wb_raw = oxl.load_workbook(INPUT_FILE, data_only=True)

def sheet_name_for(cond: str, ct: str) -> str:
    # Matches your worksheet naming convention
    return f"{cond} {ct}"

def is_number(x) -> bool:
    return isinstance(x, (int, float)) and x is not None

def norm(x) -> str:
    return str(x).strip() if x is not None else ""

# Helper: given a worksheet and a column index, walk LEFT on row 1 until a non-empty cell is found
def nearest_left_header_on_row1(ws, col_idx: int) -> str | None:
    c = col_idx
    while c >= 1:
        v = ws.cell(row=1, column=c).value
        if v not in (None, ""):
            return norm(v)
        c -= 1
    return None

# Helper: build {layer -> column_index} for the (Layer, Myelination) subcolumns.
# We detect columns where row2 == "Myelination", then infer the layer by scanning row1 left to the nearest non-empty cell.
def myelination_columns_by_layer(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        sub = norm(ws.cell(row=2, column=col).value)
        if sub == "Myelination":
            layer_label = nearest_left_header_on_row1(ws, col)
            if layer_label:
                mapping[layer_label] = col
    return mapping

# Build pooled RAW lists: (condition, layer) -> list of float
pooled_raw: dict[tuple[str, str], list[float]] = {(cond, layer): [] for cond in CONDITIONS for layer in LAYERS}

# For sanity reporting: count RAW numeric cells consumed per (cond, ct)
raw_counts_per_sheet: dict[tuple[str, str], int] = {}

# Sweep all 9 input sheets; collect RAW per-row Myelination values (no averaging)
for cond in CONDITIONS:
    for ct in CELL_TYPES:
        ws_name = sheet_name_for(cond, ct)
        if ws_name not in wb_raw.sheetnames:
            print(f"[WARN] Missing worksheet: {ws_name}")
            raw_counts_per_sheet[(cond, ct)] = 0
            continue
        ws = wb_raw[ws_name]

        # Map each layer -> the column holding (Layer, Myelination)
        layer_to_col = myelination_columns_by_layer(ws)
        # Warn if any expected layer is missing
        for expected_layer in LAYERS:
            if expected_layer not in layer_to_col:
                print(f"[WARN] {ws_name}: Missing Myelination column for layer '{expected_layer}'")

        # Data rows begin after the two header rows
        start_row = START_ROW
        end_row = ws.max_row

        # Collect RAW numeric cells
        used = 0
        for layer in LAYERS:
            col = layer_to_col.get(layer, -1)
            if col == -1:
                continue
            for r in range(start_row, end_row + 1):
                val = ws.cell(row=r, column=col).value
                if is_number(val):
                    pooled_raw[(cond, layer)].append(float(val))
                    used += 1

        raw_counts_per_sheet[(cond, ct)] = used

# 4c) Write a two-row header: Row 1 = primary (conditions), Row 2 = subcolumns (layers)
header_row_1 = 1
header_row_2 = 2
bold_font = Font(bold=True)
center_align = Alignment(horizontal="center", vertical="center")

col_idx = 1
for cond in CONDITIONS:
    start_col = col_idx
    for layer in LAYERS:
        cell = ws_all.cell(row=header_row_2, column=col_idx, value=layer)
        cell.font = bold_font
        cell.alignment = center_align
        col_idx += 1
    end_col = col_idx - 1
    ws_all.merge_cells(start_row=header_row_1, end_row=header_row_1, start_column=start_col, end_column=end_col)
    top = ws_all.cell(row=header_row_1, column=start_col, value=cond)
    top.font = bold_font
    top.alignment = center_align

# 4d) Fill values: each column holds RAW values for one (condition × layer); rows stack observations
data_start_row = 3
for cond_block_i, cond in enumerate(CONDITIONS):
    for layer_i, layer in enumerate(LAYERS):
        base = cond_block_i * len(LAYERS)  # 0, 7, 14
        col = base + layer_i + 1           # 1..21
        series_vals = pooled_raw[(cond, layer)]
        for r, val in enumerate(series_vals, start=data_start_row):
            ws_all.cell(row=r, column=col, value=val)

# Freeze header
ws_all.freeze_panes = "A3"

# 4e) Sanity check: RAW numeric cells from source vs written entries
print("—" * 60)
print("Sanity check for Prism Prep - All (RAW Myelination entries)")
total_raw_source = 0
for cond in CONDITIONS:
    for ct in CELL_TYPES:
        cnt = raw_counts_per_sheet.get((cond, ct), 0)
        total_raw_source += cnt
        print(f"{cond} — {ct}: raw entries (non-empty numeric) = {cnt}")
print(f"Total RAW entries across 9 sheets (source): {total_raw_source}")

total_written = sum(len(pooled_raw[(cond, layer)]) for cond in CONDITIONS for layer in LAYERS)
print(f"Total entries written to Prism Prep - All: {total_written}")

if total_written != total_raw_source:
    print(f"""⚠️  WARNING: Entry count mismatch!
    - RAW entries in source: {total_raw_source}
    - Written to 'Prism Prep - All': {total_written}
    This indicates an extraction/column-mapping issue (header mismatch, missing sheet/column, etc.).""")
else:
    print("✅  Entry counts match (RAW).")
print("—" * 60)

# 4f) Save the workbook
wb_out.save(out_path)
print("✅  Prism Prep - Donor sheet added.")
print(f"✅  Prism Prep - All sheet added.  Saved to: {out_path}")
