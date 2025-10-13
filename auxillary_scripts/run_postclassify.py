#!/usr/bin/env python3
"""
run_postclassify.py — Spyder-friendly standalone runner (with annotations)

What this does
--------------
- Recursively finds `Hyperspectral_Results_*.xlsx` under BASE_DIR.
- Reads peak fits, computes features, classifies (CH-stretch only).
- Writes per-file `*_classified.csv` and a "Classification" sheet back into each workbook.
- Builds a consolidated CSV at BASE_DIR that ALSO includes per-droplet annotations
  from each workbook's "Raw Data" sheet:
    * Category, Location, Cell Marker, LAMP2_Coloc (joined by DropletID/Lipid ID)
  and two regex-derived fields from the source file path:
    * clinicogenotype: one of {Control, AD33, AD44, Unknown}
    * cortical_layer: one of {L1..L6, WM, N/A}

Usage (Spyder)
--------------
1) Open this file in Spyder.
2) Confirm BASE_DIR below.
3) Run ▶

Requirements: pandas, openpyxl. This file should be in the same folder as
`classify_rules.py` and `hyperspec_features.py`, or run from the repo root
so `lipid_analysis` is importable.
"""
from __future__ import annotations

import os
import sys
import re
import glob
import traceback
from typing import List, Optional, Iterable

import pandas as pd

# -------------------------
# USER SETTINGS
# -------------------------
BASE_DIR = r"D:\OneDrive - Stanford\Research Documents\AD Project\2025"
SHEET_OVERRIDE: Optional[str] = None   # e.g., "Peak Fits" or None to auto-detect
DRY_RUN: bool = False                  # if True, don't write back to Excel
CONSOLIDATE: bool = True               # write consolidated CSV in BASE_DIR
CONSOLIDATE_NAME: str = "Hyperspectral_Classification_Summary.csv"
GLOB_PATTERN: str = "**/Hyperspectral_Results_*.xlsx"  # recursive discovery
RULES_JSON: Optional[str] = None       # path to rules override JSON
# -------------------------

# Try local imports first, fall back to package import if needed
try:
    from classify_rules import load_rules, classify_table
    from hyperspec_features import compute_features_table
except Exception:
    HERE = os.path.dirname(os.path.abspath(__file__))
    PARENT = os.path.dirname(HERE)
    if PARENT not in sys.path:
        sys.path.insert(0, PARENT)
    try:
        from lipid_analysis.classify_rules import load_rules, classify_table  # type: ignore
        from lipid_analysis.hyperspec_features import compute_features_table  # type: ignore
    except Exception as e:
        raise ImportError(
            "Could not import required modules. Ensure this script is in the same folder as "
            "classify_rules.py/hyperspec_features.py OR run it from the repo root so "
            "'lipid_analysis' is importable.\nOriginal error: " + str(e)
        )


# -------------------------
# Helpers: discovery & sheet selection
# -------------------------

def discover_all(base_dir: str, pattern: str) -> List[str]:
    """Recursively find matching Excel workbooks under base_dir."""
    search = os.path.join(base_dir, pattern)
    files = sorted(glob.glob(search, recursive=True))
    return [f for f in files if os.path.isfile(f)]


def pick_sheet(xlsx_path: str, override: Optional[str]) -> Optional[str]:
    """Choose the appropriate sheet to read for peak fits."""
    try:
        xls = pd.ExcelFile(xlsx_path, engine="openpyxl")
    except Exception:
        return None
    if override is not None and override in xls.sheet_names:
        return override
    if "Peak Fits" in xls.sheet_names:
        return "Peak Fits"
    if xls.sheet_names:
        return xls.sheet_names[min(2, len(xls.sheet_names) - 1)]
    return None


# -------------------------
# Helpers: regex extraction from source_file path
# -------------------------

def parse_clinicogenotype(path_like: str) -> str:
    # Search the full relative path for robustness
    s = path_like.lower()
    if "ad44" in s:
        return "AD44"
    if "ad33" in s:
        return "AD33"
    if "control" in s or re.search(r"(?<![a-z0-9])ctrl(?![a-z0-9])", s):
        return "Control"
    return "Unknown"


def parse_cortical_layer(path_like: str) -> str:
    # Look for standalone L1..L6 or WM tokens anywhere in the path
    m = re.search(r"(?<![A-Za-z0-9])(L[1-6]|WM)(?![A-Za-z0-9])", path_like, flags=re.IGNORECASE)
    if not m:
        return "N/A"
    token = m.group(1).upper()
    return token


# -------------------------
# Helpers: annotations from Raw Data sheet
# -------------------------
RAW_SHEET_CANDIDATES = ["Raw Data", "RawData", "RAW DATA"]
RAW_ID_COL_CANDIDATES = ["DropletID", "Droplet_ID", "Lipid ID", "LipidID", "ID"]
RAW_ANN_COLUMNS = ["Category", "Location", "Cell Marker", "LAMP2_Coloc"]


def read_raw_annotations(xlsx_path: str) -> Optional[pd.DataFrame]:
    """Read Raw Data sheet and return a (id, annotations) table.

    Will attempt to detect an ID column among RAW_ID_COL_CANDIDATES and
    keep only columns relevant for joining.
    """
    try:
        xls = pd.ExcelFile(xlsx_path, engine="openpyxl")
    except Exception:
        return None

    raw_sheet = None
    for nm in RAW_SHEET_CANDIDATES:
        if nm in xls.sheet_names:
            raw_sheet = nm
            break
    if raw_sheet is None:
        return None

    try:
        rdf = pd.read_excel(xlsx_path, sheet_name=raw_sheet, engine="openpyxl")
    except Exception:
        return None

    # Find ID column
    id_col = None
    cols_lower = {c.lower(): c for c in rdf.columns}
    for cand in RAW_ID_COL_CANDIDATES:
        key = cand.lower()
        if key in cols_lower:
            id_col = cols_lower[key]
            break
    if id_col is None:
        # heuristic: any column containing both 'id' and one of 'droplet','lipid'
        for c in rdf.columns:
            cl = c.lower()
            if "id" in cl and ("droplet" in cl or "lipid" in cl):
                id_col = c
                break
    if id_col is None:
        return None

    keep_cols = [id_col] + [c for c in RAW_ANN_COLUMNS if c in rdf.columns]
    if len(keep_cols) == 1:
        # only ID present; nothing to add
        return None

    rdf2 = rdf[keep_cols].copy()
    # Normalize ID column name for merge convenience
    rdf2.rename(columns={id_col: "__JOIN_ID__"}, inplace=True)
    return rdf2


def detect_out_id_column(df: pd.DataFrame) -> Optional[str]:
    """Try to find the droplet ID column in the classifier output table."""
    for cand in RAW_ID_COL_CANDIDATES:
        if cand in df.columns:
            return cand
    # heuristic fallback
    for c in df.columns:
        cl = c.lower()
        if "id" in cl and ("droplet" in cl or "lipid" in cl):
            return c
    return None


# -------------------------
# Core per-file classify + annotate
# -------------------------

def classify_single_file(
    filepath: str,
    rules_json: Optional[str],
    sheet_override: Optional[str],
    write_back: bool,
) -> Optional[pd.DataFrame]:
    sheet = pick_sheet(filepath, sheet_override)
    if sheet is None:
        print(f"[postclassify/spyder] Skipping (no readable sheets): {filepath}")
        return None

    try:
        df = pd.read_excel(filepath, sheet_name=sheet, engine="openpyxl")
    except Exception as e:
        print(f"[postclassify/spyder] Skipping {os.path.basename(filepath)}: failed to read '{sheet}' ({e})")
        return None

    rules = load_rules(rules_json)
    feats = compute_features_table(df)
    out = classify_table(feats, rules)

    # Join per-droplet annotations from Raw Data
    rdf = read_raw_annotations(filepath)
    out_id_col = detect_out_id_column(out)
    if rdf is not None and out_id_col is not None:
        rdf_join = rdf.copy()
        # If out uses a different ID name, align on values
        out = out.merge(
            rdf_join,
            how="left",
            left_on=out_id_col,
            right_on="__JOIN_ID__",
        )
        if "__JOIN_ID__" in out.columns:
            out.drop(columns=["__JOIN_ID__"], inplace=True)

    # Per-file CSV next to workbook
    csv_out = os.path.splitext(filepath)[0] + "_classified.csv"
    try:
        out.to_csv(csv_out, index=False)
        print(f"[postclassify/spyder] CSV: {csv_out} ({len(out)} rows)")
    except Exception as e:
        print(f"[postclassify/spyder] Failed CSV for {filepath}: {e}")

    # Write Classification sheet
    if write_back:
        try:
            with pd.ExcelWriter(
                filepath, mode="a", engine="openpyxl", if_sheet_exists="replace"
            ) as writer:
                out.to_excel(writer, sheet_name="Classification", index=False)
            print(f"[postclassify/spyder] Wrote 'Classification' sheet → {os.path.basename(filepath)}")
        except TypeError:
            from openpyxl import load_workbook
            try:
                wb = load_workbook(filepath)
                if "Classification" in wb.sheetnames:
                    ws = wb["Classification"]
                    wb.remove(ws)
                    wb.save(filepath)
                with pd.ExcelWriter(filepath, mode="a", engine="openpyxl") as writer:
                    out.to_excel(writer, sheet_name="Classification", index=False)
                print(f"[postclassify/spyder] Wrote 'Classification' sheet → {os.path.basename(filepath)}")
            except Exception as e:
                print(f"[postclassify/spyder] Failed Excel write for {filepath}: {e}")
        except Exception as e:
            print(f"[postclassify/spyder] Failed Excel write for {filepath}: {e}")

    return out


# -------------------------
# Consolidation
# -------------------------

def consolidate(rows: Iterable[pd.DataFrame], records_meta: List[dict], base_dir: str, name: str) -> None:
    """Combine all per-file tables into one CSV, adding source-derived fields."""
    dfs = []
    for r, meta in zip(rows, records_meta):
        if r is None or len(r) == 0:
            continue
        r2 = r.copy()
        # add meta columns to each chunk
        r2.insert(0, "source_file", meta["source_file"])  # relative path under BASE_DIR
        r2.insert(1, "sheet_used", meta["sheet_used"])    # actual sheet or 'auto'
        r2.insert(2, "clinicogenotype", meta["clinicogenotype"])  # Control/AD33/AD44/Unknown
        r2.insert(3, "cortical_layer", meta["cortical_layer"])    # L1..L6/WM/N/A
        dfs.append(r2)

    if not dfs:
        print("[postclassify/spyder] No rows to consolidate.")
        return

    big = pd.concat(dfs, axis=0, ignore_index=True)
    out_path = os.path.join(base_dir, name)
    try:
        big.to_csv(out_path, index=False)
        print(f"[postclassify/spyder] Consolidated CSV: {out_path} ({len(big)} rows)")
    except Exception as e:
        print(f"[postclassify/spyder] Failed to write consolidated CSV: {e}")


# -------------------------
# Main runner
# -------------------------

def run() -> None:
    print("[postclassify/spyder] Starting…")
    if not os.path.isdir(BASE_DIR):
        print(f"[postclassify/spyder] BASE_DIR does not exist: {BASE_DIR}")
        return

    files = discover_all(BASE_DIR, GLOB_PATTERN)
    if not files:
        print(f"[postclassify/spyder] No workbooks found under: {BASE_DIR}")
        return

    print(f"[postclassify/spyder] Found {len(files)} workbook(s) under {BASE_DIR}\n")

    all_rows: List[pd.DataFrame] = []
    metas: List[dict] = []
    for i, fp in enumerate(files, 1):
        rel = os.path.relpath(fp, BASE_DIR)
        print(f"[{i}/{len(files)}] {rel}")
        try:
            out = classify_single_file(
                fp,
                rules_json=RULES_JSON,
                sheet_override=SHEET_OVERRIDE,
                write_back=not DRY_RUN,
            )
            if out is not None:
                meta = {
                    "source_file": rel,
                    "sheet_used": SHEET_OVERRIDE or "auto",
                    "clinicogenotype": parse_clinicogenotype(rel),
                    "cortical_layer": parse_cortical_layer(rel),
                }
                metas.append(meta)
                all_rows.append(out)
        except KeyboardInterrupt:
            print("[postclassify/spyder] Interrupted by user.")
            return
        except Exception:
            print(f"[postclassify/spyder] Error processing {rel}:\n{traceback.format_exc()}")

    if CONSOLIDATE and all_rows:
        consolidate(all_rows, metas, BASE_DIR, CONSOLIDATE_NAME)
    else:
        print("[postclassify/spyder] Consolidation disabled or no rows produced.")

    print("\n[postclassify/spyder] Done.")


# Auto-run when executed in Spyder
run()
