import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

SUBCOLUMNS = [
    "Lipids",
    "Lipidated Lipofuscin",
    "Lipofuscin",
    "Myelination",
    "Amyloid"
]

# Only these acquisition files are eligible for cortical-layer stratification.
# All other files should NOT be layer-mapped and will be skipped.
LAYERED_FILES = {
    # Microglia / Astrocytes — S0536
    "Control-S0536-DAPI-IBA1-GFAP-100X-StacksMicroglia.nd2",
    "Control-S0536-DAPI-IBA1-GFAP-100X-StacksAstrocytes.nd2",

    # AD33 S2143
    "AD33-S2143-DAPI-IBA1-GFAP-100X-StacksMicroglia.nd2",
    "AD33-S2143-DAPI-IBA1-GFAP-100X-StacksAstrocytes.nd2",

    # AD44 S1342
    "AD44-S1342-DAPI-IBA1-GFAP-100X-StacksMicroglia.nd2",
    "AD44-S1342-DAPI-IBA1-GFAP-100X-StacksAstrocytes.nd2",

    # Control TUJ/LAMP2 S0536
    "Control-S0536-DAPI-TUJ-LAMP2-100X-StacksNeurons.nd2",

    # AD33 S2146
    "AD33-S2146-DAPI-IBA1-GFAP-100X-StacksMicroglia.nd2",
    "AD33-S2146-DAPI-IBA1-GFAP-100X-StacksAstrocytes.nd2",

    # AD44 S1563
    "AD44-S1563-DAPI-IBA1-GFAP-100X-StacksMicroglia.nd2",
    "AD44-S1563-DAPI-IBA1-GFAP-100X-StacksAstrocytes.nd2",

    # Control S2302
    "Control-S2302-DAPI-IBA1-GFAP-100X-StacksMicroglia.nd2",
    "Control-S2302-DAPI-IBA1-GFAP-100X-StacksAstrocytes.nd2",

    # TUJ/LAMP2 sets with normal mapping
    "AD33-S2146-DAPI-TUJ-LAMP2-100X-StacksNeurons.nd2",
    "AD44-S1563-DAPI-TUJ-LAMP2-100X-StacksNeurons.nd2",
    "Control-S2302-DAPI-TUJ-LAMP2-100X-StacksNeurons.nd2",
}

def determine_layer(file_name, z_stack):
    """
    Map z-slice index to cortical layer using a default 6-slices-per-layer bucketing,
    but ONLY for allowed acquisition files. Three files have special offsets
    (skip early layers). All other files are not layer-stratified and are skipped.
    Returns 1..7 (Layer I..VI, White Matter) or None if not applicable/invalid.
    """
    # Normalize to basename to avoid path mismatches
    base = os.path.basename(str(file_name)).strip()

    # Special files that skip early layers (your original exceptions)
    OFFSETS = {
        "AD44-S1342-DAPI-TUJ-LAMP2-100X-StacksNeurons.nd2": 1,  # skip L1
        "AD33-S2143-DAPI-TUJ-LAMP2-100X-StacksNeurons.nd2": 2,  # skip L1,L2
        "Control-S2218-DAPI-TUJ-LAMP2-100X-StacksNeurons.nd2": 3,  # skip L1–L3
    }

    # Only proceed if this file is meant to be layer-mapped.
    # Allowed if in the allowlist, or one of the special-offset files.
    if base not in LAYERED_FILES and base not in OFFSETS:
        return None

    # Default: 6 z-slices per layer ⇒ 1..6→L1, 7..12→L2, ..., 37..→L7 (WM)
    def base_layer(z):
        try:
            z_int = int(z)
        except (TypeError, ValueError):
            return None
        if z_int <= 0:
            return None
        L = 1 + (z_int - 1) // 6
        return min(L, 7)

    L0 = base_layer(z_stack)
    if L0 is None:
        return None

    k = OFFSETS.get(base, 0)
    L = L0 + k
    return L if 1 <= L <= 7 else None


def create_formatted_worksheet(wb, sheet_name):
    """
    Create a worksheet with merged headers:
      Row 1 => "file_name" (col A) and then each layer name spanning its subcolumns
      Row 2 => subcolumn names from SUBCOLUMNS
    """
    ws = wb.create_sheet(title=sheet_name)

    # Column A: file_name header merged over two rows
    ws.cell(row=1, column=1, value="file_name")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=1, column=1).font = Font(bold=True)

    # Layer headers and subcolumns
    layer_names = ["Layer I", "Layer II", "Layer III", "Layer IV", "Layer V", "Layer VI", "White Matter"]
    subcolumns = SUBCOLUMNS
    n_sub = len(subcolumns)

    # Start writing layers in column B
    start_col = 2
    for layer_name in layer_names:
        # Merge row 1 across n_sub columns for this layer
        ws.merge_cells(
            start_row=1,
            start_column=start_col,
            end_row=1,
            end_column=start_col + n_sub - 1
        )
        # Write the layer title
        cell = ws.cell(row=1, column=start_col, value=layer_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write each subcolumn in row 2
        for j, subcol_name in enumerate(subcolumns):
            c = start_col + j
            cell = ws.cell(row=2, column=c, value=subcol_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Advance by n_sub columns for the next layer
        start_col += n_sub

    return ws

def main():
    input_filename = r"D:\OneDrive - Stanford\Research Documents\AD Project\2025\AD Lipid Statistics.xlsx"
    output_filename = r"D:\OneDrive - Stanford\Research Documents\AD Project\2025\AD_Lipid_Statistics_CorticalLayers.xlsx"

    xl = pd.ExcelFile(input_filename)
    sheet_names = xl.sheet_names

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name in sheet_names:
        df = pd.read_excel(xl, sheet_name=sheet_name)

        # Check columns
        required_cols = [
            'file_name',
            'z_stack',
            'pure_lipid_percentage',
            'lipid_lipofuscin_percentage',
            'lipofuscin_percentage',
            'myelination_percentage',
            'amyloid_percentage'
        ]
        if not all(col in df.columns for col in required_cols):
            print(f"Skipping '{sheet_name}' - missing required columns.")
            continue

        # Create the new sheet with the 2-row headers
        ws = create_formatted_worksheet(wb, sheet_name)

        # Group the data by file_name -> layer -> list of (lipids, lipid_lipo, lipo)
        grouped_data = {}
        for i, row_data in df.iterrows():
            fn_raw = row_data['file_name']
            fn = str(fn_raw).strip()                   # ← normalize once, up front
            zs = row_data['z_stack']
            pl = row_data['pure_lipid_percentage']
            llf = row_data['lipid_lipofuscin_percentage']
            lpf = row_data['lipofuscin_percentage']
            mb  = row_data['myelination_percentage']
            ap  = row_data['amyloid_percentage']
            
            layer = determine_layer(fn, zs)            # ← use normalized fn
            if layer is None:
                print(f"[WARN] Skipping row (no layer): file='{fn}', z_stack={zs}")
                continue
            
            # Store in grouped_data
            if fn not in grouped_data:
                grouped_data[fn] = {i: [] for i in range(1,8)}
            grouped_data[fn][layer].append((int(zs) if pd.notna(zs) else None, pl, llf, lpf, mb, ap))

        # Now we’ll write the data. Start from row=3 (below the headers).
        current_row = 3

        # We have 7 layer blocks (Layer I..WM),
        # each taking up len(SUBCOLUMNS)=5 subcolumns => total columns from 2..36
        def layer_to_cols(layer):
            # layer 1 => columns 2,3,4
            # layer 2 => columns 5,6,7
            # ...
            start_col = 2 + (layer - 1)*len(SUBCOLUMNS)
            return tuple(start_col + i for i in range(len(SUBCOLUMNS)))

        # Write each file_name in a block
        for file_name, layer_dict in grouped_data.items():
            # Sort within each layer by z_stack for stable row order
            for lyr in range(1, 8):
                layer_vals = layer_dict[lyr]
                # keep tuples as (z, pl, llf, lpf, mb, ap)
                layer_vals.sort(key=lambda tup: (tup[0] is None, tup[0]))

            # For each of the 7 layers, we have some number of data points
            # The block height = the maximum number of data points among all layers
            max_points = max(len(layer_dict[layer]) for layer in range(1, 8))

            # If a file_name has 0 data points altogether, skip
            if max_points == 0:
                continue

            # We will fill rows current_row..(current_row+max_points-1)
            # and place data in the correct columns for each layer
            block_start = current_row
            block_end   = current_row + max_points - 1

            # (Optional) Put the file_name in col=1, merged from block_start..block_end
            ws.cell(row=block_start, column=1, value=file_name)
            if max_points > 1:
                # vertically merge that cell down
                ws.merge_cells(
                    start_row=block_start,
                    start_column=1,
                    end_row=block_end,
                    end_column=1
                )
            # style it
            ws.cell(row=block_start, column=1).alignment = Alignment(
                horizontal='center',
                vertical='center'
            )
            ws.cell(row=block_start, column=1).font = Font(bold=False)

            # Now fill row by row
            # For each i in 0..(max_points-1), place data for each layer in that row
            for i in range(max_points):
                row_i = block_start + i

                for lyr in range(1, 8):
                    data_points = grouped_data[file_name][lyr]
                    cols = layer_to_cols(lyr)  # now returns 5 column indices
                    if i < len(data_points):
                        # data_points[i] is a 5-tuple: (pl, llf, lpf, mb, ap)
                        vals = data_points[i][1:]  # drop z
                        for val, col in zip(vals, cols):
                            ws.cell(row=row_i, column=col, value=val)

            # After filling these max_points rows for this file_name,
            # move current_row down for the next file_name block
            current_row = block_end + 1

    # Finally, save the new workbook
    wb.save(output_filename)
    print(f"Saved '{output_filename}' with your desired layout.")

if __name__ == "__main__":
    main()
