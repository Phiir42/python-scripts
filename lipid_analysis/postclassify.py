# lipid_analysis/postclassify.py
"""
Post-processing utilities to classify objects in Hyperspectral_Results_*.xlsx files
using CH-stretch rules (no fingerprint required).

- Scans a data directory for result workbooks.
- Reads the appropriate sheet (prefers "Peak Fits") from each workbook.
- Computes features and applies rule-based classification.
- Writes a "Classification" sheet back into each workbook (optional).
- Emits per-file CSVs and a consolidated summary CSV (optional).
"""

from __future__ import annotations

import glob
import logging
import os
from typing import List, Optional

import pandas as pd

from .classify_rules import classify_table, load_rules
from .constants import LOG_LEVEL
from .hyperspec_features import compute_features_table

logger = logging.getLogger(__name__)
logger.setLevel(LOG_LEVEL)


def classify_hyperspectral_dir(
    directory: str,
    rules_json: Optional[str] = None,
    write_back: bool = True,
    consolidate: bool = True,
) -> str:
    """
    Classify all Hyperspectral_Results_*.xlsx workbooks in `directory`.

    For each workbook:
      1) Choose the "Peak Fits" sheet if present; otherwise, fall back to a later sheet.
      2) Compute feature table and apply rule-based classification.
      3) Save a per-file CSV: <workbook_stem>_classified.csv.
      4) Optionally write a "Classification" worksheet back into the workbook.

    Optionally, also write a consolidated CSV aggregating all files.

    Parameters
    ----------
    directory : str
        Directory to scan for Hyperspectral_Results_*.xlsx files.
    rules_json : Optional[str]
        Path to a rules JSON (if None, uses built-in defaults in `load_rules`).
    write_back : bool
        If True, writes a "Classification" sheet into each workbook.
    consolidate : bool
        If True, writes a consolidated CSV across all processed files.

    Returns
    -------
    str
        Path to the consolidated CSV if created; otherwise an empty string.
    """
    pattern = os.path.join(directory, "Hyperspectral_Results_*.xlsx")
    files = sorted(glob.glob(pattern))
    rules = load_rules(rules_json)

    consolidated_rows: List[pd.DataFrame] = []

    for workbook_path in files:
        try:
            xls = pd.ExcelFile(workbook_path, engine="openpyxl")
            # Prefer "Peak Fits"; otherwise pick a later sheet defensively
            chosen_sheet = (
                "Peak Fits"
                if "Peak Fits" in xls.sheet_names
                else xls.sheet_names[min(2, max(0, len(xls.sheet_names) - 1))]
            )
            df = pd.read_excel(xls, sheet_name=chosen_sheet)
        except Exception as exc:
            logger.warning(
                "[postclassify] Skipping %s: failed to read sheet (%s)",
                workbook_path,
                exc,
            )
            continue

        feats = compute_features_table(df)
        classified = classify_table(feats, rules)

        # Per-file CSV
        csv_out = os.path.splitext(workbook_path)[0] + "_classified.csv"
        try:
            classified.to_csv(csv_out, index=False)
        except Exception as exc:
            logger.warning(
                "[postclassify] Failed writing per-file CSV %s (%s)", csv_out, exc
            )

        # Optional write-back to the workbook (new/replace 'Classification' sheet)
        if write_back:
            try:
                with pd.ExcelWriter(
                    workbook_path,
                    mode="a",
                    engine="openpyxl",
                    if_sheet_exists="replace",
                ) as writer:
                    classified.to_excel(writer, sheet_name="Classification", index=False)
            except TypeError:
                # Older pandas: no if_sheet_exists — remove sheet manually if present
                try:
                    from openpyxl import load_workbook  # type: ignore

                    wb = load_workbook(workbook_path)
                    if "Classification" in wb.sheetnames:
                        ws = wb["Classification"]
                        wb.remove(ws)
                        wb.save(workbook_path)
                    with pd.ExcelWriter(
                        workbook_path, mode="a", engine="openpyxl"
                    ) as writer:
                        classified.to_excel(
                            writer, sheet_name="Classification", index=False
                        )
                except Exception as exc:
                    logger.warning(
                        "[postclassify] Failed to write 'Classification' sheet to %s (%s)",
                        workbook_path,
                        exc,
                    )
            except Exception as exc:
                logger.warning(
                    "[postclassify] Failed to write 'Classification' sheet to %s (%s)",
                    workbook_path,
                    exc,
                )

        # Gather for consolidation
        annotated = classified.copy()
        annotated.insert(0, "source_file", os.path.basename(workbook_path))
        annotated.insert(1, "sheet_used", chosen_sheet)
        consolidated_rows.append(annotated)

    consolidated_path = ""
    if consolidate and consolidated_rows:
        try:
            big = pd.concat(consolidated_rows, axis=0, ignore_index=True)
            consolidated_path = os.path.join(
                directory, "Hyperspectral_Classification_Summary.csv"
            )
            big.to_csv(consolidated_path, index=False)
            logger.info(
                "[postclassify] Wrote consolidated summary: %s (%d rows)",
                consolidated_path,
                len(big),
            )
        except Exception as exc:
            logger.warning(
                "[postclassify] Failed to write consolidated summary (%s)", exc
            )
    else:
        logger.info("[postclassify] No hyperspectral results found to classify.")

    return consolidated_path
